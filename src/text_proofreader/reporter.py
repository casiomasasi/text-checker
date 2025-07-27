"""
レポート生成機能

校正結果をExcel形式で出力し、統計情報や分析結果を提供する。
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Union
from collections import Counter


class Reporter:
    """校正結果レポートを生成するクラス"""
    
    def __init__(self):
        """Reporterを初期化"""
        self.results = []
        self.statistics = {}
        self.metadata = {}
        
    def add_results(self, results: List[Dict], metadata: Dict = None):
        """
        校正結果を追加
        
        Args:
            results: 校正結果のリスト
            metadata: ファイルのメタデータ
        """
        self.results.extend(results)
        if metadata:
            self.metadata.update(metadata)
        
        # 統計情報を更新
        self._update_statistics()
    
    def _update_statistics(self):
        """統計情報を更新"""
        if not self.results:
            return
        
        # エラータイプ別集計
        error_types = [result.get('type', 'unknown') for result in self.results]
        type_counts = Counter(error_types)
        
        # 重要度別集計
        severities = [result.get('severity', 'unknown') for result in self.results]
        severity_counts = Counter(severities)
        
        # サブタイプ別集計
        subtypes = [result.get('subtype', 'unknown') for result in self.results]
        subtype_counts = Counter(subtypes)
        
        self.statistics = {
            'total_errors': len(self.results),
            'error_types': dict(type_counts),
            'severities': dict(severity_counts),
            'subtypes': dict(subtype_counts),
            'file_info': self.metadata
        }
    
    def generate_excel_report(self, output_path: Union[str, Path]) -> str:
        """
        Excel形式のレポートを生成
        
        Args:
            output_path: 出力ファイルパス
            
        Returns:
            生成されたファイルのパス
        """
        output_path = Path(output_path)
        
        # 複数シートのExcelファイルを作成
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # メインの校正結果シート
            self._create_main_sheet(writer)
            
            # 統計情報シート
            self._create_statistics_sheet(writer)
            
            # エラータイプ別シート
            self._create_error_type_sheets(writer)
            
            # ファイル情報シート
            self._create_file_info_sheet(writer)
        
        return str(output_path)
    
    def _create_main_sheet(self, writer):
        """メインの校正結果シートを作成"""
        if not self.results:
            pd.DataFrame({'メッセージ': ['校正結果がありません']}).to_excel(
                writer, sheet_name='校正結果', index=False
            )
            return
        
        # データフレーム用のデータを準備
        data = []
        for i, result in enumerate(self.results, 1):
            data.append({
                'No.': i,
                '行番号': result.get('line', ''),
                '列番号': result.get('column', ''),
                'エラータイプ': result.get('type', ''),
                'サブタイプ': result.get('subtype', ''),
                '重要度': result.get('severity', ''),
                '検出箇所': result.get('original', ''),
                '修正案': result.get('correction', result.get('suggestion', '')),
                '説明': result.get('description', ''),
                '例': ', '.join(result.get('examples', [])) if result.get('examples') else ''
            })
        
        df = pd.DataFrame(data)
        
        # Excelに書き込み
        df.to_excel(writer, sheet_name='校正結果', index=False)
        
        # 書式設定
        worksheet = writer.sheets['校正結果']
        
        # 列幅を調整
        column_widths = {
            'A': 8,   # No.
            'B': 10,  # 行番号
            'C': 10,  # 列番号
            'D': 20,  # エラータイプ
            'E': 15,  # サブタイプ
            'F': 10,  # 重要度
            'G': 25,  # 検出箇所
            'H': 25,  # 修正案
            'I': 40,  # 説明
            'J': 30   # 例
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        
        # 重要度に応じた色付け
        self._apply_severity_colors(worksheet, df)
    
    def _apply_severity_colors(self, worksheet, df):
        """重要度に応じて行に色を付ける"""
        from openpyxl.styles import PatternFill
        
        # 色の定義
        colors = {
            'high': PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid'),    # 薄い赤
            'medium': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'), # 薄い黄
            'low': PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')     # 薄い緑
        }
        
        # データ行に色を適用（ヘッダー行は除く）
        for row_idx, severity in enumerate(df['重要度'], start=2):
            if severity in colors:
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = colors[severity]
    
    def _create_statistics_sheet(self, writer):
        """統計情報シートを作成"""
        stats_data = []
        
        # 基本統計
        stats_data.append(['項目', '値'])
        stats_data.append(['総エラー数', self.statistics.get('total_errors', 0)])
        stats_data.append(['', ''])
        
        # エラータイプ別統計
        stats_data.append(['エラータイプ別集計', ''])
        error_types = self.statistics.get('error_types', {})
        for error_type, count in error_types.items():
            stats_data.append([error_type, count])
        
        stats_data.append(['', ''])
        
        # 重要度別統計
        stats_data.append(['重要度別集計', ''])
        severities = self.statistics.get('severities', {})
        for severity, count in severities.items():
            stats_data.append([severity, count])
        
        df_stats = pd.DataFrame(stats_data)
        df_stats.to_excel(writer, sheet_name='統計情報', header=False, index=False)
        
        # 列幅調整
        worksheet = writer.sheets['統計情報']
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 15
    
    def _create_error_type_sheets(self, writer):
        """エラータイプ別の詳細シートを作成"""
        error_types = self.statistics.get('error_types', {})
        
        for error_type in error_types.keys():
            # 該当するエラーのみを抽出
            filtered_results = [r for r in self.results if r.get('type') == error_type]
            
            if not filtered_results:
                continue
            
            # データフレーム作成
            data = []
            for i, result in enumerate(filtered_results, 1):
                data.append({
                    'No.': i,
                    '行番号': result.get('line', ''),
                    '列番号': result.get('column', ''),
                    'サブタイプ': result.get('subtype', ''),
                    '重要度': result.get('severity', ''),
                    '検出箇所': result.get('original', ''),
                    '修正案': result.get('correction', result.get('suggestion', '')),
                    '説明': result.get('description', '')
                })
            
            df = pd.DataFrame(data)
            
            # シート名を作成（Excelの制限に対応）
            sheet_name = self._create_valid_sheet_name(error_type)
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 列幅調整
            worksheet = writer.sheets[sheet_name]
            column_widths = {'A': 8, 'B': 10, 'C': 10, 'D': 15, 'E': 10, 'F': 25, 'G': 25, 'H': 40}
            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width
    
    def _create_file_info_sheet(self, writer):
        """ファイル情報シートを作成"""
        file_info = []
        
        file_info.append(['項目', '値'])
        file_info.append(['ファイル名', self.metadata.get('file_name', '')])
        file_info.append(['ファイルパス', self.metadata.get('file_path', '')])
        file_info.append(['ファイルサイズ', f"{self.metadata.get('file_size', 0)} bytes"])
        file_info.append(['エンコーディング', self.metadata.get('encoding', '')])
        file_info.append(['行数', self.metadata.get('line_count', 0)])
        file_info.append(['文字数', self.metadata.get('char_count', 0)])
        file_info.append(['拡張子', self.metadata.get('extension', '')])
        file_info.append(['処理日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        df_info = pd.DataFrame(file_info)
        df_info.to_excel(writer, sheet_name='ファイル情報', header=False, index=False)
        
        # 列幅調整
        worksheet = writer.sheets['ファイル情報']
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 50
    
    def _create_valid_sheet_name(self, name: str) -> str:
        """
        Excelで有効なシート名を作成
        
        Args:
            name: 元の名前
            
        Returns:
            有効なシート名
        """
        # Excelで使用できない文字を置換
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        valid_name = name
        
        for char in invalid_chars:
            valid_name = valid_name.replace(char, '_')
        
        # 長さ制限（31文字まで）
        if len(valid_name) > 31:
            valid_name = valid_name[:28] + '...'
        
        return valid_name
    
    def get_summary(self) -> Dict:
        """
        校正結果のサマリーを取得
        
        Returns:
            サマリー情報の辞書
        """
        if not self.results:
            return {'message': '校正結果がありません'}
        
        # 重要度別の評価スコア計算
        severity_weights = {'high': 3, 'medium': 2, 'low': 1}
        total_weight = sum(severity_weights.get(r.get('severity', 'low'), 1) for r in self.results)
        
        # ファイルサイズに基づくスコア正規化
        char_count = self.metadata.get('char_count', 1)
        error_density = len(self.results) / max(char_count / 1000, 1)  # 1000文字あたりのエラー数
        
        # 評価スコア（100点満点、エラーが少ないほど高得点）
        max_score = 100
        penalty = min(total_weight * 2 + error_density * 10, max_score)
        score = max(0, max_score - penalty)
        
        return {
            'total_errors': len(self.results),
            'error_types': self.statistics.get('error_types', {}),
            'severities': self.statistics.get('severities', {}),
            'error_density': round(error_density, 2),
            'evaluation_score': round(score, 1),
            'file_info': {
                'name': self.metadata.get('file_name', ''),
                'size': self.metadata.get('file_size', 0),
                'char_count': self.metadata.get('char_count', 0)
            }
        }
    
    def print_summary(self):
        """サマリーをコンソールに出力"""
        summary = self.get_summary()
        
        if 'message' in summary:
            print(summary['message'])
            return
        
        print("=== 校正結果サマリー ===")
        print(f"ファイル名: {summary['file_info']['name']}")
        print(f"文字数: {summary['file_info']['char_count']:,} 文字")
        print(f"総エラー数: {summary['total_errors']} 件")
        print(f"エラー密度: {summary['error_density']} 件/1000文字")
        print(f"評価スコア: {summary['evaluation_score']}/100")
        print()
        
        print("エラータイプ別:")
        for error_type, count in summary['error_types'].items():
            print(f"  {error_type}: {count} 件")
        print()
        
        print("重要度別:")
        for severity, count in summary['severities'].items():
            print(f"  {severity}: {count} 件")
    
    def clear_results(self):
        """結果をクリア"""
        self.results = []
        self.statistics = {}
        self.metadata = {}